import streamlit as st
import openpyxl
from datetime import date

#Date
today = date.today()

# UI for the page
st.markdown("<h1 style='text-align: center; color: black;'>Daily Tracker ++</h1>", unsafe_allow_html=True)


def save_in_daily_tracker(sessions, learnings, project_tasks):
    path = "tracker.xlsx"
    
    wb_obj = openpyxl.load_workbook(path)  
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column

    filled_row = sheet_obj.max_row 
    sheet_obj.cell(row = 1, column = 1, value = "Date")
    sheet_obj.cell(row = 1, column = 2, value = "Session details")
    sheet_obj.cell(row = 1, column = 3, value = "Learning details")
    sheet_obj.cell(row = 1, column = 4, value = "Project related tasks")
        
    sheet_obj.cell(row=filled_row + 1, column=1, value = today)
    sheet_obj.cell(row=filled_row + 1, column=2, value = sessions)
    sheet_obj.cell(row=filled_row + 1, column=3, value = learnings)
    sheet_obj.cell(row=filled_row + 1, column=4, value = project_tasks)
    
    wb_obj.save(path)


sessions = st.text_area("Tell about the sessions you attended today", "Type Here ...")
learnings = st.text_area("Tell about the learnings for the day", "Type Here ...")
project_tasks = st.text_area("Tell about your project related tasks", "Type Here ...")


if(st.button('Submit')):
    save_in_daily_tracker(sessions.title(), learnings.title(), project_tasks.title())
    st.success("Hurray! Saved successfully")
    st.balloons()
